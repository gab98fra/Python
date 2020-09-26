# ----------------------------------------------------------------------------
# Nombre:       excel.py
# Autor:        Gabriel F
# GitHub:       https://github.com/gab98fra/
# Creado:       22 de Septiembre 2020
# Modificado:   25 de Septiembre 2020
# Copyright:    (c) 2020 by Gabriel F, 2020
# ----------------------------------------------------------------------------

"""
    Permite leer, escribir y exportar datos archivo excel utilizando Programación Funcional

    Python 3.8.2
    openpyxl 3.0.5

"""

from openpyxl import load_workbook, Workbook
import datetime


def read_excel():
#Leer archivos excel con rango expecificado
    
    #Leer archivo excel
    book=load_workbook("data.xlsx", data_only=True)
    
    #Activar hoja
    sheet1=book.active

    #Indicar rango
    range=sheet1["A2":"D11"]

    #Almacenará todos los datos
    all_data=[]

    for row in range:
    #Recorremos las filas

        #obtemos valores por cada fila
        data=[range.value for range in row]
        
        #Agregar a la variable
        all_data.append(data)
    
    for row in all_data:
    #Imprimir los datos por fila

        print (f'Datos de la fila {row[0]} ;\n {row} \n')


def read_excel_all():
#Leer todos los datos de excel sin especificar el rango

    #Leer archivo excel
    book=load_workbook("data.xlsx", data_only=True)

    #Activar hoja
    sheet1=book.active

    #Última fila/columna con información
    #max_num_row=sheet1.max_row
    max_num_column=sheet1.max_column
    
    for row in sheet1.iter_rows(min_col=1, max_col=max_num_column):
    #recorrer todas las filas

        #obtener el dato por cada fila
        data=[cell.value for cell in row]
       
        print (data)


def write_excel():
#Guardar archivos excel

    #Crear archivo excel
    book=Workbook()

    #Acivar hoja
    sheet1=book.active

    #Asignar valores a la hoja
    sheet1["A1"]="Hola amigo curioso"
    sheet1["A2"]=f'Hoy es: {datetime.datetime.now()}, y hoy tendrá un gra día'
    sheet1["A3"]="Este es un ejemplo de como crear y escribir archivos excel"

    sheet1.append([1,2,3])
    sheet1.append(["col1", "col2", "col3"])

    #Guardar archivo
    book.save('file_excel.xlsx')


if __name__ == "__main__":
    
    #read_excel()
    #write_excel()
    read_excel_all()